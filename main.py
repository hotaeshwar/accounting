import random
import string
from fastapi import FastAPI, Depends, HTTPException, Response, status, Header, Request
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Union
from enum import Enum
from datetime import datetime, timedelta
import sqlite3
import os
import jwt
import pandas as pd
import uuid
import calendar
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.styles import PatternFill, Font
from io import BytesIO

# Constants
DATABASE_NAME = "expense_tracker.db"
SECRET_KEY = "09d25e094faa6ca2556c818166b7a9563b93f7099f6f0f4caa6cf63b88e8d3e7"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
EXCEL_EXPORT_DIRECTORY = "exports"
ADMIN_SECRET_KEY = "734bb15a7c471de7b40bebe1b0dad8fefc05654984f36411d7b79ebbe7e9df77"  # Change this in production
ARCHIVE_RETENTION_DAYS = 30
# Create exports directory if it doesn't exist
if not os.path.exists(EXCEL_EXPORT_DIRECTORY):
    os.makedirs(EXCEL_EXPORT_DIRECTORY)

# Initialize FastAPI app
app = FastAPI(title="Office Expense Tracker API",
              description="API for tracking office expenses with admin and guest user roles")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://accountsonline.info", "https://www.accountsonline.info", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# Define expense categories
class ExpenseCategory(str, Enum):
    OFFICE_SUPPLIES = "Office Supplies"
    RENT = "Rent"
    UTILITIES = "Utilities"
    SALARY = "Salary"
    TRAVEL = "Travel"
    EQUIPMENT = "Equipment"
    SOFTWARE = "Software"
    MARKETING = "Marketing"
    INSURANCE = "Insurance"
    TAXES = "Taxes"
    MAINTENANCE = "Maintenance"
    TELECOMMUNICATIONS = "Telecommunications"
    CONSULTING = "Consulting"
    LEGAL = "Legal"
    MISCELLANEOUS = "Miscellaneous"

# Define user role
class UserRole(str, Enum):
    ADMIN = "admin"
    GUEST = "guest"

# Database setup with trigger
def init_db():
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    # Users table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL
    )
    ''')

    # Add admin restriction trigger
    cursor.execute('''
    CREATE TRIGGER IF NOT EXISTS prevent_multi_admin
    BEFORE INSERT ON users
    FOR EACH ROW
    WHEN NEW.role = 'admin' AND EXISTS (SELECT 1 FROM users WHERE role = 'admin')
    BEGIN
        SELECT RAISE(ABORT, 'Only one admin allowed');
    END;
    ''')

    # Expenses table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id TEXT UNIQUE NOT NULL,
        user_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        category TEXT NOT NULL,
        description TEXT,
        date_created TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )
    ''')

    # Income table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS income (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        description TEXT,
        amount REAL NOT NULL,
        date_created TEXT NOT NULL
    )
    ''')

    conn.commit()
    conn.close()

# Initialize database
init_db()
# Pydantic models
class UserCreate(BaseModel):
    username: str
    password: str
    role: UserRole
    admin_key: Optional[str] = None  # For admin registration only

class User(BaseModel):
    id: int
    username: str
    role: UserRole

class LoginRequest(BaseModel):
    username: str
    password: str
    role: Optional[UserRole] = None

class TokenData(BaseModel):
    access_token: str
    token_type: str
    user_role: str

class TokenResponse(BaseModel):
    success: bool = True
    data: TokenData
    message: str = "Login successful"

class PasswordResetRequest(BaseModel):
    username: str

class PasswordResetVerify(BaseModel):
    username: str
    reset_code: str
    new_password: str

class ExpenseBase(BaseModel):
    amount: float
    category: ExpenseCategory
    description: Optional[str] = None

class ExpenseCreate(ExpenseBase):
    pass

class Expense(ExpenseBase):
    id: int
    invoice_id: str
    user_id: int
    date_created: str

class ResponseModel(BaseModel):
    success: bool
    data: Optional[Any] = None
    message: Optional[str] = None
    admin_exists: Optional[bool] = None  # Added for frontend control

class CategoryItem(BaseModel):
    key: str
    name: str

class CategoriesResponse(BaseModel):
    success: bool = True
    data: List[CategoryItem]
    message: str = "Categories retrieved successfully"

class IncomeCreate(BaseModel):
    description: str
    amount: float

class ProfitLossData(BaseModel):
    total_income: float
    total_expenses: float
    net_profit_loss: float
    expenses_by_category: Dict[str, float]

# User authentication functions
def get_user(username: str):
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, password, role FROM users WHERE username = ?", (username,))
    user_data = cursor.fetchone()
    conn.close()

    if user_data:
        return {"id": user_data[0], "username": user_data[1], "password": user_data[2], "role": user_data[3]}
    return None

def admin_exists():
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (UserRole.ADMIN,))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0

def authenticate_user(username: str, password: str, role: Optional[UserRole] = None):
    user = get_user(username)
    if not user or user["password"] != password:
        return False

    if role and user["role"] != role:
        return False

    return user

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(
    authorization: Optional[str] = Header(None),
    x_auth_token: Optional[str] = Header(None)
):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )

    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.split("Bearer ")[1]
    elif x_auth_token:
        if x_auth_token.startswith("Bearer "):
            token = x_auth_token.split("Bearer ")[1]
        else:
            token = x_auth_token

    if not token:
        raise credentials_exception

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        user_id: int = payload.get("user_id")
        role: str = payload.get("role")

        if username is None or user_id is None:
            raise credentials_exception

    except jwt.PyJWTError as e:
        print(f"JWT decode error: {str(e)}")
        raise credentials_exception

    user = get_user(username)
    if user is None:
        raise credentials_exception

    return user

async def get_current_admin(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    return current_user

# Helper functions
def generate_invoice_id():
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    unique_id = str(uuid.uuid4()).split('-')[0]
    return f"INV-{timestamp}-{unique_id}"

password_reset_codes = {}

def generate_reset_code(length=6):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))

def purge_expired_archives():
    """Purge archives that have exceeded their retention period"""
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Delete expired archived expenses
    cursor.execute(
        "DELETE FROM archived_expenses WHERE archive_expiry < ?", 
        (current_datetime,)
    )
    
    # Delete expired archived income
    cursor.execute(
        "DELETE FROM archived_income WHERE archive_expiry < ?", 
        (current_datetime,)
    )
    
    conn.commit()
    conn.close()

def export_expenses_to_excel_stream(year, month):
    """
    Generate an Excel file with expenses and income data with totals
    Returns:
        tuple: (excel_binary, filename, total_expenses, total_income, net_amount)
    """
    try:
        conn = sqlite3.connect(DATABASE_NAME)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Format dates for query
        start_date = f"{year}-{month:02d}-01"
        _, last_day = calendar.monthrange(year, month)
        end_date = f"{year}-{month:02d}-{last_day} 23:59:59"

        # Get expenses data - including user role information
        cursor.execute('''
            SELECT e.id, e.invoice_id, e.amount, e.category, e.description, e.date_created,
                   u.username, u.role
            FROM expenses e
            JOIN users u ON e.user_id = u.id
            WHERE e.date_created BETWEEN ? AND ?
            ORDER BY e.date_created
        ''', (start_date, end_date))
        expenses = [dict(row) for row in cursor.fetchall()]

        # Get income data
        cursor.execute('''
            SELECT id, description, amount, date_created
            FROM income
            WHERE date_created BETWEEN ? AND ?
            ORDER BY date_created
        ''', (start_date, end_date))
        income = [dict(row) for row in cursor.fetchall()]

        # Calculate totals
        total_expenses = sum(exp['amount'] for exp in expenses) if expenses else 0
        total_income = sum(inc['amount'] for inc in income) if income else 0
        net_amount = total_income - total_expenses

        # Create timestamp for unique filename
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"financial_report_{year}_{month:02d}_{timestamp}.xlsx"
        
        # Create Excel file in memory
        excel_binary = BytesIO()
        writer = pd.ExcelWriter(excel_binary, engine='openpyxl')
        workbook = writer.book

        # Add expenses sheet with total
        if expenses:
            df_expenses = pd.DataFrame(expenses)
            df_expenses.to_excel(writer, index=False, sheet_name='Expenses')
            ws_exp = writer.sheets['Expenses']
            
            # Formatting - Highlight guest entries with blue background, admin with red
            for idx, row in enumerate(df_expenses.itertuples(index=False), start=2):
                cell_color = "E6F2FF" if row.role == "guest" else "FFE6E6"  # Blue for guest, Red for admin
                for col in range(1, len(df_expenses.columns) + 1):
                    ws_exp.cell(row=idx, column=col).fill = PatternFill(
                        start_color=cell_color, end_color=cell_color, fill_type="solid"
                    )
            
            # Add totals row at the bottom
            total_row = len(df_expenses) + 3
            ws_exp.cell(row=total_row, column=1, value="Total Expenses").font = Font(bold=True)
            ws_exp.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})").font = Font(bold=True)
            
            # Apply bold to total and formatting
            ws_exp.cell(row=total_row, column=3).fill = PatternFill(
                start_color="FFD700", end_color="FFD700", fill_type="solid"
            )

        # Add income sheet with total
        if income:
            df_income = pd.DataFrame(income)
            df_income.to_excel(writer, index=False, sheet_name='Income')
            ws_inc = writer.sheets['Income']
            
            # Add totals row at the bottom
            total_row = len(df_income) + 3
            ws_inc.cell(row=total_row, column=1, value="Total Income").font = Font(bold=True)
            ws_inc.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})").font = Font(bold=True)
            
            # Apply bold to total and formatting
            ws_inc.cell(row=total_row, column=3).fill = PatternFill(
                start_color="98FB98", end_color="98FB98", fill_type="solid"
            )

        # Add summary sheet with more detailed breakdown
        if expenses or income:
            # Create category breakdown
            categories = {}
            for exp in expenses:
                category = exp['category']
                if category not in categories:
                    categories[category] = 0
                categories[category] += exp['amount']
            
            # Summary data with categories
            summary_data = {
                'Metric': ['Total Income', 'Total Expenses', 'Net Profit/Loss'] + list(categories.keys()),
                'Amount': [total_income, total_expenses, net_amount] + list(categories.values())
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, index=False, sheet_name='Summary')
            ws_sum = writer.sheets['Summary']
            
            # Format net profit/loss
            ws_sum.cell(row=4, column=2).fill = PatternFill(
                start_color="98FB98" if net_amount >= 0 else "FF6347",
                end_color="98FB98" if net_amount >= 0 else "FF6347",
                fill_type="solid"
            )
            
            # Format headers and make them bold
            for cell in ws_sum[1]:
                cell.font = Font(bold=True)
            
            # Format the category breakdown section header
            if categories:
                ws_sum.cell(row=4, column=1, value="Category Breakdown:").font = Font(bold=True)

        # Auto-adjust column widths
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 5

        writer.close()
        excel_binary.seek(0)
        
        return excel_binary.getvalue(), filename, total_expenses, total_income, net_amount
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return None, None, 0, 0, 0
    finally:
        conn.close()
def archive_financial_data(year, month):
    """Archive both expenses and income for the given month/year"""
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    start_date = f"{year}-{month:02d}-01"
    _, last_day = calendar.monthrange(year, month)
    end_date = f"{year}-{month:02d}-{last_day} 23:59:59"
    
    archive_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Set expiry date to 30 days from now
    archive_expiry = (datetime.now() + timedelta(days=ARCHIVE_RETENTION_DAYS)).strftime("%Y-%m-%d %H:%M:%S")

    # Archive expenses
    cursor.execute('''
        INSERT INTO archived_expenses
        (invoice_id, user_id, amount, category, description, date_created, archive_date, archive_expiry)
        SELECT invoice_id, user_id, amount, category, description, date_created, ?, ?
        FROM expenses
        WHERE date_created BETWEEN ? AND ?
    ''', (archive_date, archive_expiry, start_date, end_date))

    cursor.execute('''
        DELETE FROM expenses WHERE date_created BETWEEN ? AND ?
    ''', (start_date, end_date))

    # Archive income
    cursor.execute('''
        INSERT INTO archived_income
        (description, amount, date_created, archive_date, archive_expiry)
        SELECT description, amount, date_created, ?, ?
        FROM income
        WHERE date_created BETWEEN ? AND ?
    ''', (archive_date, archive_expiry, start_date, end_date))

    cursor.execute('''
        DELETE FROM income WHERE date_created BETWEEN ? AND ?
    ''', (start_date, end_date))

    conn.commit()
    conn.close()

# API endpoints
@app.post("/register", response_model=ResponseModel)
async def register(user: UserCreate):
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    
    try:
        # Check if username exists
        cursor.execute("SELECT id FROM users WHERE username = ?", (user.username,))
        if cursor.fetchone():
            conn.close()
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": "Username already registered"}
            )

        # Admin registration logic
        if user.role == UserRole.ADMIN:
            # Check if admin exists (atomic operation)
            cursor.execute("BEGIN IMMEDIATE")
            cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (UserRole.ADMIN,))
            admin_count = cursor.fetchone()[0]
            
            if admin_count > 0:
                conn.rollback()
                conn.close()
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={
                        "success": False,
                        "message": "Admin registration is disabled",
                        "admin_exists": True
                    }
                )

            # Verify admin key
            if user.admin_key != ADMIN_SECRET_KEY:
                conn.rollback()
                conn.close()
                return JSONResponse(
                    status_code=status.HTTP_403_FORBIDDEN,
                    content={"success": False, "message": "Invalid admin registration key"}
                )

        # Insert the new user
        cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                      (user.username, user.password, user.role))
        user_id = cursor.lastrowid
        conn.commit()

        user_data = {"id": user_id, "username": user.username, "role": user.role}
        return {
            "success": True,
            "data": user_data,
            "message": "User registered successfully",
            "admin_exists": user.role == UserRole.ADMIN
        }
    except sqlite3.IntegrityError as e:
        if "Only one admin allowed" in str(e):
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={
                    "success": False,
                    "message": "Admin user already exists. Only one admin is allowed.",
                    "admin_exists": True
                }
            )
        conn.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    except Exception as e:
        conn.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()

@app.post("/login", response_model=TokenResponse)
async def login(login_data: LoginRequest):
    user = authenticate_user(login_data.username, login_data.password, login_data.role)
    if not user:
        return JSONResponse(
            status_code=status.HTTP_401_UNAUTHORIZED,
            content={
                "success": False,
                "message": "Invalid credentials or role"
            }
        )

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"], "user_id": user["id"], "role": user["role"]},
        expires_delta=access_token_expires
    )

    token_data = {
        "access_token": access_token,
        "token_type": "bearer",
        "user_role": user["role"]
    }

    return {
        "success": True,
        "data": token_data,
        "message": "Login successful"
    }

@app.post("/password/reset-request", response_model=ResponseModel)
async def request_password_reset(reset_request: PasswordResetRequest):
    user = get_user(reset_request.username)
    if not user:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "User not found"}
        )
    
    reset_code = generate_reset_code()
    expiration_time = datetime.utcnow() + timedelta(minutes=30)
    password_reset_codes[reset_request.username] = {
        "code": reset_code,
        "expires_at": expiration_time
    }
    
    return {
        "success": True,
        "data": {"reset_code": reset_code},
        "message": "Password reset code generated."
    }

@app.post("/password/reset-verify", response_model=ResponseModel)
async def verify_reset_and_change_password(reset_verify: PasswordResetVerify):
    user = get_user(reset_verify.username)
    if not user:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "User not found"}
        )
    
    if reset_verify.username not in password_reset_codes:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "No reset code requested for this user or code has expired"}
        )
    
    reset_data = password_reset_codes[reset_verify.username]
    
    if datetime.utcnow() > reset_data["expires_at"]:
        del password_reset_codes[reset_verify.username]
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "Reset code has expired"}
        )
    
    if reset_data["code"] != reset_verify.reset_code:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "Invalid reset code"}
        )
    
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE users SET password = ? WHERE username = ?",
            (reset_verify.new_password, reset_verify.username)
        )
        conn.commit()
        
        del password_reset_codes[reset_verify.username]
        
        return {
            "success": True,
            "message": "Password changed successfully"
        }
    except Exception as e:
        conn.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()

@app.get("/categories", response_model=CategoriesResponse)
async def get_categories():
    categories = [{"key": category.name, "name": category.value} for category in ExpenseCategory]
    return {
        "success": True,
        "data": categories,
        "message": "Categories retrieved successfully"
    }

@app.post("/expenses", response_model=ResponseModel)
async def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    invoice_id = generate_invoice_id()
    date_created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO expenses (invoice_id, user_id, amount, category, description, date_created) VALUES (?, ?, ?, ?, ?, ?)",
            (invoice_id, current_user["id"], expense.amount, expense.category, expense.description, date_created)
        )
        expense_id = cursor.lastrowid
        conn.commit()

        expense_data = {
            "id": expense_id,
            "invoice_id": invoice_id,
            "user_id": current_user["id"],
            "amount": expense.amount,
            "category": expense.category,
            "description": expense.description,
            "date_created": date_created
        }

        return {
            "success": True,
            "data": expense_data,
            "message": f"Expense created successfully with invoice ID: {invoice_id}"
        }
    except Exception as e:
        conn.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()

@app.get("/expenses/my", response_model=ResponseModel)
async def list_my_expenses(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        cursor.execute(
            "SELECT id, invoice_id, user_id, amount, category, description, date_created FROM expenses WHERE user_id = ?",
            (current_user["id"],)
        )
        expenses = []
        for row in cursor.fetchall():
            expenses.append(dict(row))

        return {"success": True, "data": expenses}
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()

@app.get("/expenses/all", response_model=ResponseModel)
async def list_all_expenses(current_user: dict = Depends(get_current_admin)):
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        # Fetch expenses with user role information for color coding
        cursor.execute(
            """
            SELECT e.id, e.invoice_id, e.user_id, e.amount, e.category, e.description, 
                   e.date_created, u.username, u.role 
            FROM expenses e 
            JOIN users u ON e.user_id = u.id
            """
        )
        expenses = []
        for row in cursor.fetchall():
            expenses.append(dict(row))

        return {"success": True, "data": expenses}
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()

@app.get("/export/current-month")
async def export_current_month_expenses(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        return JSONResponse(
            status_code=status.HTTP_403_FORBIDDEN,
            content={"success": False, "message": "Not enough permissions"}
        )

    try:
        current_date = datetime.now()
        # Fix: Unpack all 5 values or use _ for ones you don't need
        excel_data, filename, _, _, _ = export_expenses_to_excel_stream(current_date.year, current_date.month)
        
        if excel_data is None:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "message": "Failed to generate Excel file"}
            )
        
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Export error: {str(e)}"}
        )
@app.get("/invoice/{invoice_id}", response_model=ResponseModel)
async def get_invoice_by_id(invoice_id: str, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        cursor.execute(
            "SELECT e.*, u.username, u.role FROM expenses e JOIN users u ON e.user_id = u.id WHERE e.invoice_id = ?",
            (invoice_id,)
        )
        expense = cursor.fetchone()

        if not expense:
            cursor.execute(
                "SELECT * FROM archived_expenses WHERE invoice_id = ?",
                (invoice_id,)
            )
            expense = cursor.fetchone()
            if not expense:
                conn.close()
                return JSONResponse(
                    status_code=status.HTTP_404_NOT_FOUND,
                    content={"success": False, "message": "Invoice not found"}
                )

        expense_dict = dict(expense)

        if current_user["role"] != UserRole.ADMIN and expense_dict["user_id"] != current_user["id"]:
            conn.close()
            return JSONResponse(
                status_code=status.HTTP_403_FORBIDDEN,
                content={"success": False, "message": "Not authorized to view this invoice"}
            )

        return {"success": True, "data": expense_dict}
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()

@app.post("/archive/{year}/{month}")
async def archive_month_expenses(year: int, month: int, current_user: dict = Depends(get_current_admin)):
    try:
        if month < 1 or month > 12:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": "Invalid month. Must be between 1-12"}
            )

        if year < 2000 or year > datetime.now().year:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": f"Invalid year. Must be between 2000-{datetime.now().year}"}
            )

        excel_data, filename = export_expenses_to_excel_stream(year, month)
        
        if excel_data is None:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "message": "Failed to generate Excel file"}
            )
        
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Archive operation failed: {str(e)}"}
        )

@app.get("/archive/{month}/{year}", response_model=ResponseModel)
async def get_archived_expenses(month: int, year: int, current_user: dict = Depends(get_current_admin)):
    try:
        if month < 1 or month > 12:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": "Invalid month. Must be between 1-12"}
            )

        if year < 2000 or year > datetime.now().year:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": f"Invalid year. Must be between 2000-{datetime.now().year}"}
            )

        start_date = f"{year}-{month:02d}-01"
        _, last_day = calendar.monthrange(year, month)
        end_date = f"{year}-{month:02d}-{last_day} 23:59:59"

        conn = sqlite3.connect(DATABASE_NAME)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT ar.id, ar.invoice_id, ar.user_id, ar.amount, ar.category,
                   ar.description, ar.date_created, ar.archive_date, u.username
            FROM archived_expenses ar
            JOIN users u ON ar.user_id = u.id
            WHERE ar.date_created BETWEEN ? AND ?
            ORDER BY ar.date_created
            """,
            (start_date, end_date)
        )

        archived_expenses = [dict(row) for row in cursor.fetchall()]
        conn.close()

        if not archived_expenses:
            return {
                "success": True,
                "data": [],
                "message": f"No archived expenses found for {year}-{month:02d}"
            }

        return {
            "success": True,
            "data": archived_expenses,
            "message": f"Retrieved {len(archived_expenses)} archived expenses for {year}-{month:02d}"
        }
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"This feature is reserved for admin login personnel only"}
        )

@app.post("/income", response_model=ResponseModel)
async def create_income(income: IncomeCreate, current_user: dict = Depends(get_current_admin)):
    date_created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO income (description, amount, date_created) VALUES (?, ?, ?)",
            (income.description, income.amount, date_created)
        )
        income_id = cursor.lastrowid
        conn.commit()

        income_data = {
            "id": income_id,
            "description": income.description,
            "amount": income.amount,
            "date_created": date_created
        }

        return {
            "success": True,
            "data": income_data,
            "message": "Income created successfully"
        }
    except Exception as e:
        conn.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )
    finally:
        conn.close()
@app.get("/profit-loss/current-month", response_model=ResponseModel)
async def get_current_month_profit_loss(current_user: dict = Depends(get_current_user)):
    now = datetime.now()
    start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    end_date = now.strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    try:
        cursor.execute(
            "SELECT SUM(amount) FROM income WHERE date_created BETWEEN ? AND ?",
            (start_date, end_date)
        )
        total_income = cursor.fetchone()[0] or 0.0

        cursor.execute(
            "SELECT SUM(amount) FROM expenses WHERE date_created BETWEEN ? AND ?",
            (start_date, end_date)
        )
        total_expenses = cursor.fetchone()[0] or 0.0

        expenses_by_category = {}
        for category in ExpenseCategory:
            cursor.execute(
                "SELECT SUM(amount) FROM expenses WHERE category = ? AND date_created BETWEEN ? AND ?",
                (category.value, start_date, end_date)
            )
            category_total = cursor.fetchone()[0] or 0.0
            expenses_by_category[category.value] = category_total

        net_profit_loss = total_income - total_expenses

        profit_loss_data = {
            "total_income": total_income,
            "total_expenses": total_expenses,
            "net_profit_loss": net_profit_loss,
            "expenses_by_category": expenses_by_category
        }

        return {
            "success": True,
            "data": profit_loss_data,
            "message": "Profit/Loss data retrieved successfully"
        }
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error retrieving profit/loss data: {str(e)}"}
        )
    finally:
        conn.close()
# Health check endpoint
@app.get("/health")
async def health_check():
    return {"status": "healthy", "success": True}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)